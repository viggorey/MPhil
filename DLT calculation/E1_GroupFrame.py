import os
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
from scipy import stats
from scipy.stats import f_oneway
import statsmodels.api as sm
from statsmodels.stats.anova import anova_lm
from statsmodels.formula.api import ols
import warnings
warnings.filterwarnings('ignore')

# ===== CONFIGURATION =====
# Path to folder containing metadata files
DATA_FOLDER = "/Users/viggorey/Desktop/PhD/Cambridge/Macaranga/3D transformation/5. Datasets/3D data/Large branch"

# Frame rate for time conversion
FRAME_RATE = 91  # frames per second

# Groups to compare (separated by semicolons, each group can be multiple conditions separated by commas)
# Examples:
#   "11U,21U" - Compare wax runners vs non-wax runners on waxy stems going up
#   "11U,21U;11D,21D" - Compare both up and down movements
#   "11U,12U" - Compare wax runners on waxy vs non-waxy stems going up
COMPARISON_GROUPS = "11U,12U,21U,22U"

# Sheet and column to compare (format: "SheetName:ColumnName") 
# Examples:
#   "Body_Measurements:Middle_Leg_Length_Avg_Normalized"
#   "Kinematics_Speed:Speed_ThoraxLengths_per_s"
#   "Kinematics_Gait:Duty_Factor_Overall"
#   "Kinematics_Range:Step_Length_Front_Avg_Normalized"
#   "Body_Positioning:CoM_Overall_Branch_Distance_Normalized"
#   "Body_Positioning:gaster_left_right_angle_abs"
#   "Biomechanics:Minimum_Pull_Off_Force"
#   "Behavioral:Slip_Score"
#   "CoM:CoM_Thorax_Branch_Distance_Normalized"
METRIC = "Kinematics_Range:Tibia_Stem_Angle_Hind_Avg"

# List of filenames to highlight in yellow (without the 'meta_' prefix and '.xlsx' extension)
# Example: ["11U1", "21U2"]
HIGHLIGHTED_SAMPLES = [""]

# Color scheme for groups (can be modified)
COLOR_SCHEME = {
    '11U': '#1f77b4',  # blue
    '21U': '#ff7f0e',  # orange
    '12U': '#2ca02c',  # green
    '22U': '#d62728',  # red
    '11D': '#9467bd',  # purple
    '21D': '#8c564b',  # brown
    '12D': '#e377c2',  # pink
    '22D': '#7f7f7f'   # gray
}

# Base output folder (relative to script location)
OUTPUT_FOLDER = "comparison_plots"

# Prefix mapping for sheet names
SHEET_PREFIXES = {
    'Kinematics': 'KN',
    'CoM': 'CoM',
    'Coordinate_System': 'CS',
    'Behavioral_Scores': 'BS',
    '3D_Coordinates': '3D',
    'Duty_factor': 'DF',
}

COMPARISON_EXCEL_DIR = "/Users/viggorey/Desktop/PhD/Cambridge/Macaranga/3D transformation/5. Datasets/3D data/Large branch/Comparisons"
# =========================

def sanitize_filename(filename):
    """Replace problematic characters in filenames with underscores."""
    # Replace forward slashes, backslashes, and other problematic characters
    problematic_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    for char in problematic_chars:
        filename = filename.replace(char, '_')
    return filename

def load_metadata_file(file_path):
    """Load all sheets from a metadata Excel file."""
    try:
        return pd.read_excel(file_path, sheet_name=None)
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return None

def parse_group_name(filename):
    """Extract group information from filename (e.g., trim_meta_11U2.xlsx -> 11U)."""
    base_name = os.path.basename(filename)
    if base_name.startswith('trim_meta_'):
        return base_name[10:13]  # Extract the group code (e.g., 11U)
    return None

def get_metric_data(metadata, sheet_name, column_name):
    """Extract specific metric data from metadata sheets."""
    if sheet_name in metadata and column_name in metadata[sheet_name].columns:
        return metadata[sheet_name][column_name]
    return None

def parse_factorial_design(group_name):
    """
    Parse group name to extract factorial design factors.
    
    Args:
        group_name: String like "11U", "21D", etc.
    
    Returns:
        dict: Dictionary with 'runner_type' and 'stem_type' factors
    """
    if len(group_name) >= 3:
        first_digit = group_name[0]  # 1 or 2
        second_digit = group_name[1]  # 1 or 2
        direction = group_name[2]     # U or D
        
        # Define factors based on your experimental design
        if first_digit == '1':
            runner_type = 'Wax Runner'
        elif first_digit == '2':
            runner_type = 'Non-wax Runner'
        else:
            runner_type = 'Unknown'
            
        if second_digit == '1':
            stem_type = 'Waxy Stem'
        elif second_digit == '2':
            stem_type = 'Non-waxy Stem'
        else:
            stem_type = 'Unknown'
            
        return {
            'runner_type': runner_type,
            'stem_type': stem_type,
            'direction': direction
        }
    return None

def prepare_anova_data(data_dict):
    """
    Prepare data for 2x2 ANOVA analysis.
    
    Args:
        data_dict: Dictionary with group data
    
    Returns:
        DataFrame: Long-format data ready for ANOVA
    """
    anova_data = []
    
    for group, group_data in data_dict.items():
        factorial_info = parse_factorial_design(group)
        if factorial_info:
            for file_name, data in group_data.items():
                # Calculate mean across time for each individual
                mean_value = np.nanmean(data)
                if not np.isnan(mean_value):
                    anova_data.append({
                        'individual': file_name,
                        'group': group,
                        'runner_type': factorial_info['runner_type'],
                        'stem_type': factorial_info['stem_type'],
                        'direction': factorial_info['direction'],
                        'value': mean_value
                    })
    
    return pd.DataFrame(anova_data)

def perform_2x2_anova(anova_df):
    """
    Perform 2x2 ANOVA with main effects and interaction.
    
    Args:
        anova_df: DataFrame prepared for ANOVA
    
    Returns:
        dict: ANOVA results including F-statistics and p-values
    """
    if len(anova_df) < 4:  # Need at least 4 data points for 2x2 ANOVA
        return None
    
    try:
        # Create the model formula for 2x2 ANOVA
        model = ols('value ~ C(runner_type) + C(stem_type) + C(runner_type):C(stem_type)', data=anova_df).fit()
        
        # Perform ANOVA
        anova_table = anova_lm(model, typ=2)
        
        # Extract results
        results = {}
        for factor in anova_table.index:
            if factor != 'Residual':
                results[factor] = {
                    'F_statistic': anova_table.loc[factor, 'F'],
                    'p_value': anova_table.loc[factor, 'PR(>F)'],
                    'df_factor': anova_table.loc[factor, 'df'],
                    'df_residual': anova_table.loc['Residual', 'df']
                }
        
        return results
        
    except Exception as e:
        print(f"Error in ANOVA: {e}")
        return None

def create_anova_annotation(anova_results):
    """
    Create annotation text for ANOVA results.
    
    Args:
        anova_results: Results from perform_2x2_anova
    
    Returns:
        str: Formatted annotation text
    """
    if not anova_results:
        return "Insufficient data for ANOVA"
    
    annotation_lines = ["2x2 ANOVA Results:"]
    
    # Main effects
    if 'C(runner_type)' in anova_results:
        runner_result = anova_results['C(runner_type)']
        runner_sig = '*' if runner_result['p_value'] < 0.05 else ''
        annotation_lines.append(f"Runner Type: F({runner_result['df_factor']:.0f},{runner_result['df_residual']:.0f}) = {runner_result['F_statistic']:.3f}, p = {runner_result['p_value']:.3g}{runner_sig}")
    
    if 'C(stem_type)' in anova_results:
        stem_result = anova_results['C(stem_type)']
        stem_sig = '*' if stem_result['p_value'] < 0.05 else ''
        annotation_lines.append(f"Stem Type: F({stem_result['df_factor']:.0f},{stem_result['df_residual']:.0f}) = {stem_result['F_statistic']:.3f}, p = {stem_result['p_value']:.3g}{stem_sig}")
    
    # Interaction effect
    if 'C(runner_type):C(stem_type)' in anova_results:
        interaction_result = anova_results['C(runner_type):C(stem_type)']
        interaction_sig = '*' if interaction_result['p_value'] < 0.05 else ''
        annotation_lines.append(f"Interaction: F({interaction_result['df_factor']:.0f},{interaction_result['df_residual']:.0f}) = {interaction_result['F_statistic']:.3f}, p = {interaction_result['p_value']:.3g}{interaction_sig}")
    
    return '\n'.join(annotation_lines)

def create_factorial_summary(data_dict):
    """
    Create a summary of the factorial design structure.
    
    Args:
        data_dict: Dictionary with group data
    
    Returns:
        str: Formatted summary of experimental design
    """
    summary_lines = ["Experimental Design Summary:"]
    
    # Count individuals per group
    for group in sorted(data_dict.keys()):
        factorial_info = parse_factorial_design(group)
        if factorial_info:
            n_individuals = len(data_dict[group])
            summary_lines.append(f"{group}: {factorial_info['runner_type']} on {factorial_info['stem_type']} ({n_individuals} individuals)")
    
    # Show factorial structure
    summary_lines.append("")
    summary_lines.append("2x2 Factorial Design:")
    summary_lines.append("                Wax Runners    Non-wax Runners")
    summary_lines.append("Waxy Stems        11U, 21U        12U, 22U")
    summary_lines.append("Non-waxy Stems    11D, 21D        12D, 22D")
    
    return '\n'.join(summary_lines)

def create_comparison_plot(data_dict, metric_name, output_path):
    """Create a figure with three subplots: individual lines, mean with SD, and boxplot with significance."""
    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(12, 15), height_ratios=[1, 1, 1])
    
    # Plot individual lines
    for group, group_data in data_dict.items():
        color = COLOR_SCHEME.get(group, '#000000')  # Default to black if color not specified
        for file_name, data in group_data.items():
            # Convert frames to time (seconds)
            time_points = np.arange(len(data)) / FRAME_RATE
            # Check if this sample should be highlighted
            sample_name = file_name[10:-5]  # Remove 'trim_meta_' prefix and '.xlsx' extension
            if sample_name in HIGHLIGHTED_SAMPLES:
                ax1.plot(time_points, data, color='yellow', alpha=0.7, label=f'{group} - {file_name} (highlighted)')
            else:
                ax1.plot(time_points, data, color=color, alpha=0.3)
    
    # Add color legend for groups
    for group in data_dict.keys():
        color = COLOR_SCHEME.get(group, '#000000')
        ax1.plot([], [], color=color, label=f'{group} group', linewidth=2)
    
    # Plot mean with SD
    for group, group_data in data_dict.items():
        color = COLOR_SCHEME.get(group, '#000000')
        # Convert all series to same length (pad with NaN)
        max_length = max(len(data) for data in group_data.values())
        
        # Handle padding more robustly
        padded_data = []
        for data in group_data.values():
            # Ensure data is float type to handle NaN values
            data_float = np.array(data, dtype=float)
            # Pad to max length
            if len(data_float) < max_length:
                padding_needed = max_length - len(data_float)
                padded = np.concatenate([data_float, np.full(padding_needed, np.nan)])
            else:
                padded = data_float
            padded_data.append(padded)
        
        padded_data = np.array(padded_data)
        
        mean_data = np.nanmean(padded_data, axis=0)
        std_data = np.nanstd(padded_data, axis=0)
        
        # Convert frames to time (seconds)
        x = np.arange(len(mean_data)) / FRAME_RATE
        ax2.plot(x, mean_data, color=color, label=group, linewidth=2)
        ax2.fill_between(x, mean_data - std_data, mean_data + std_data, 
                        color=color, alpha=0.2)
    
    # Customize plots
    for ax in [ax1, ax2]:
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.set_xlabel('Time (seconds)')
        ax.set_ylabel(metric_name)
    
    ax1.set_title('Individual Traces')
    ax2.set_title('Mean ± Standard Deviation')
    
    # Add legend to both plots
    ax1.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    ax2.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    
    # --- Boxplot subplot ---
    # Gather per-individual means for each group
    boxplot_data = []
    boxplot_groups = []
    highlighted_indices = []  # Store indices of highlighted samples
    for group, group_data in data_dict.items():
        for file_name, data in group_data.items():
            mean_val = np.nanmean(data)
            if not np.isnan(mean_val):  # Only include non-NaN means
                boxplot_data.append(mean_val)
                boxplot_groups.append(group)
                # Check if this sample should be highlighted
                sample_name = file_name[10:-5]  # Remove 'trim_meta_' prefix and '.xlsx' extension
                if sample_name in HIGHLIGHTED_SAMPLES:
                    highlighted_indices.append(len(boxplot_data) - 1)
    
    if boxplot_data:  # Only create boxplot if we have data
        boxplot_df = pd.DataFrame({
            'Value': boxplot_data,
            'Group': boxplot_groups
        })
        boxplot_df = boxplot_df.dropna()

        # Create boxplot with hue parameter to avoid warning
        sns.boxplot(x='Group', y='Value', data=boxplot_df, ax=ax3, palette=COLOR_SCHEME, legend=False)
        
        # Add highlighted points
        if highlighted_indices:
            highlighted_values = [boxplot_data[i] for i in highlighted_indices]
            highlighted_groups = [boxplot_groups[i] for i in highlighted_indices]
            # Get x positions for highlighted points
            x_positions = []
            for group in highlighted_groups:
                # Get the index of the group in the unique groups list
                x_pos = list(boxplot_df['Group'].unique()).index(group)
                x_positions.append(x_pos)
            
            # Add some random jitter to x positions to avoid overlap
            x_positions = [x + np.random.uniform(-0.2, 0.2) for x in x_positions]
            
            # Plot highlighted points
            ax3.scatter(x_positions, highlighted_values, color='yellow', edgecolor='black', s=100, zorder=5)
        
        ax3.set_title('Per-Individual Mean Values by Group (Boxplot)')
        ax3.set_ylabel(metric_name)
        ax3.set_xlabel('Group')

        # --- Statistical test and annotation ---
        # Prepare data for 2x2 ANOVA
        anova_df = prepare_anova_data(data_dict)
        
        if not anova_df.empty and len(anova_df['group'].unique()) >= 2:
            # Perform 2x2 ANOVA
            anova_results = perform_2x2_anova(anova_df)
            
            if anova_results:
                # Create ANOVA annotation
                anova_text = create_anova_annotation(anova_results)
                
                # Place annotation below the boxplot
                y_min = boxplot_df['Value'].min()
                y_max = boxplot_df['Value'].max()
                x_center = len(boxplot_df['Group'].unique()) / 2 - 0.5
                
                # Split text into lines for better formatting
                lines = anova_text.split('\n')
                for i, line in enumerate(lines):
                    y_offset = y_min - (0.15 + i * 0.05) * abs(y_max - y_min)
                    ax3.text(x_center, y_offset, line, ha='center', va='top', fontsize=10, 
                            bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
            else:
                # Fallback to simple tests if ANOVA fails
                unique_groups = sorted(boxplot_df['Group'].unique())
                if len(unique_groups) >= 2:
                    group_values = [boxplot_df[boxplot_df['Group'] == g]['Value'].values for g in unique_groups]
                    if len(unique_groups) == 2:
                        # Independent t-test
                        stat, pval = stats.ttest_ind(group_values[0], group_values[1], nan_policy='omit')
                        test_name = 't-test'
                    else:
                        # One-way ANOVA
                        stat, pval = stats.f_oneway(*group_values)
                        test_name = 'ANOVA'

                    # Annotate p-value
                    if pval is not None:
                        signif = '*' if pval < 0.05 else ''
                        pval_text = f"p = {pval:.3g}{signif} ({test_name})"
                        # Place annotation below the boxplot
                        y_min = boxplot_df['Value'].min()
                        x_center = len(unique_groups) / 2 - 0.5
                        y_offset = y_min - 0.1 * abs(y_max - y_min)
                        ax3.text(x_center, y_offset, pval_text, ha='center', va='top', fontsize=12,
                                bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))

    ax3.grid(True, linestyle='--', alpha=0.7)

    plt.tight_layout()
    
    # Ensure the output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    plt.savefig(output_path, bbox_inches='tight', dpi=300)
    plt.close()

def export_to_excel(data_dict, sheet_name, column_name, comparison_groups):
    # Determine prefix
    prefix = SHEET_PREFIXES.get(sheet_name, sheet_name[:2])
    excel_sheet_name = f"{prefix}_{column_name}"
    # Remove invalid Excel sheet name characters
    for char in '/\\?*[]:':
        excel_sheet_name = excel_sheet_name.replace(char, '')

    # Sort group names for file naming
    sorted_groups = sorted(comparison_groups)
    excel_file_name = "_".join(sorted_groups) + ".xlsx"
    excel_path = os.path.join(COMPARISON_EXCEL_DIR, excel_file_name)
    os.makedirs(COMPARISON_EXCEL_DIR, exist_ok=True)

    # Prepare DataFrame
    # Find max length for padding
    max_length = 0
    valid_data = []
    for group, group_data in data_dict.items():
        for file_name, data in group_data.items():
            if len(data) > 0 and not np.all(np.isnan(data)):  # Only include non-empty, non-all-NaN data
                max_length = max(max_length, len(data))
                valid_data.append((group, file_name, data))
    
    if not valid_data:  # If no valid data, don't create Excel file
        return
        
    columns = ['FileName', 'Group'] + [f'Time_{i/FRAME_RATE:.3f}s' for i in range(max_length)]
    rows = []
    for group, file_name, data in valid_data:
        padded = list(data) + [float('nan')] * (max_length - len(data))
        rows.append([file_name, group] + padded)
    
    # Create DataFrame with explicit dtypes
    df = pd.DataFrame(rows, columns=columns)
    df = df.astype({
        'FileName': str,
        'Group': str,
        **{f'Time_{i/FRAME_RATE:.3f}s': float for i in range(max_length)}
    })

    # Add mean and std rows for each group, with an empty row in between
    new_rows = []
    for group in sorted(data_dict.keys()):
        group_df = df[df['Group'] == group]
        if not group_df.empty:  # Only process groups with data
            # Mean and std rows
            group_rows = group_df.iloc[:, 2:].to_numpy(dtype=float)
            if len(group_rows) > 0:  # Check if we have data to process
                # Use masked arrays to handle NaN values
                masked_data = np.ma.masked_invalid(group_rows)
                if masked_data.count() > 0:  # Only calculate if we have valid data
                    mean_row = [f'{group} MEAN', group] + list(np.ma.mean(masked_data, axis=0).filled(np.nan))
                    std_row = [f'{group} STD', group] + list(np.ma.std(masked_data, axis=0).filled(np.nan))
                    mean_df = pd.DataFrame([mean_row], columns=df.columns)
                    std_df = pd.DataFrame([std_row], columns=df.columns)
                    # Ensure consistent dtypes
                    mean_df = mean_df.astype(df.dtypes)
                    std_df = std_df.astype(df.dtypes)
                    # Append group data, mean, std, then empty row
                    new_rows.append(group_df)
                    new_rows.append(mean_df)
                    new_rows.append(std_df)
                    empty_row = pd.DataFrame([[None] * len(df.columns)], columns=df.columns)
                    empty_row = empty_row.astype(df.dtypes)
                    new_rows.append(empty_row)
    
    if new_rows:  # Only create Excel if we have data
        # Filter out any empty DataFrames before concatenation
        new_rows = [df for df in new_rows if not df.empty]
        if new_rows:  # Double check we still have data after filtering
            # Concatenate with explicit dtypes
            df_final = pd.concat(new_rows, ignore_index=True)
            df_final = df_final.astype(df.dtypes)

            # Write to Excel (overwrite sheet if exists)
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
            else:
                wb = Workbook()
                # Remove default sheet if present
                if 'Sheet' in wb.sheetnames:
                    std = wb['Sheet']
                    wb.remove(std)
            # Remove sheet if it exists
            if excel_sheet_name in wb.sheetnames:
                ws = wb[excel_sheet_name]
                wb.remove(ws)
            ws = wb.create_sheet(excel_sheet_name)
            for r in dataframe_to_rows(df_final, index=False, header=True):
                ws.append(r)
            # Sort sheets alphabetically
            sheetnames = sorted(wb.sheetnames)
            wb._sheets = [wb[n] for n in sheetnames]
            wb.save(excel_path)

def main():
    # Get the script's directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Parse sheet and column name
    sheet_name, column_name = METRIC.split(':')
    
    # Create output folder structure (relative to script location)
    safe_column_name = sanitize_filename(column_name)
    metric_folder = os.path.join(script_dir, OUTPUT_FOLDER, sheet_name, safe_column_name)
    os.makedirs(metric_folder, exist_ok=True)
    
    # Find all metadata files
    metadata_files = [f for f in os.listdir(DATA_FOLDER) if f.startswith('trim_meta_') and f.endswith('.xlsx')]
    
    # Parse comparison groups
    comparison_groups = [group.split(',') for group in COMPARISON_GROUPS.split(';')]
    
    # Load all metadata files
    metadata_dict = {}
    for file in metadata_files:
        group = parse_group_name(file)
        if group:
            file_path = os.path.join(DATA_FOLDER, file)
            metadata_dict[group] = load_metadata_file(file_path)
    
    # Process each comparison
    for comparison in comparison_groups:
        # Collect data for the metric
        data_dict = {}
        for group in comparison:
            if group in metadata_dict:
                data_dict[group] = {}
                for file in metadata_files:
                    if file.startswith(f'trim_meta_{group}'):
                        file_path = os.path.join(DATA_FOLDER, file)
                        metadata = load_metadata_file(file_path)
                        if metadata:
                            data = get_metric_data(metadata, sheet_name, column_name)
                            if data is not None:
                                data_dict[group][file] = data
        
        if data_dict:
            # Print factorial design summary
            factorial_summary = create_factorial_summary(data_dict)
            print(f"\n{factorial_summary}")
            
            # Create plot
            plot_name = f"{'_'.join(comparison)}_{safe_column_name}.png"
            output_path = os.path.join(metric_folder, plot_name)
            create_comparison_plot(data_dict, column_name, output_path)
            print(f"Created plot: {output_path}")
            # Export to Excel
            export_to_excel(data_dict, sheet_name, column_name, comparison)

if __name__ == "__main__":
    main()
