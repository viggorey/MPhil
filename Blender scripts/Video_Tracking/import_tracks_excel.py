import bpy
import os
from openpyxl import load_workbook
from bpy.props import StringProperty, PointerProperty
from bpy.types import PropertyGroup

bl_info = {
    "name": "Import Video Track from Excel",
    "description": "Imports tracked xy coordinates from an Excel file",
    "author": "Lorenz Mammen",
    "blender": (2, 80, 0),
    "category": "Tracking",
    "support": "COMMUNITY"
}

def import_tracks_from_excel(file_path, clip):
    # Ensure we have a valid file path
    if not file_path or not os.path.isfile(file_path):
        raise ValueError("Invalid Excel file path")
        
    # Load the Excel workbook
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Get video dimensions
    video_width = clip.size[0]
    video_height = clip.size[1]
    
    # Read headers to get frame numbers
    headers = [cell.value for cell in sheet[1]][2:]  # Skip first two columns (Track Name and Coordinate Type)
    frame_numbers = [int(header.split()[1]) for header in headers]
    
    # Dictionary to store track data
    tracks_data = {}
    current_track = None
    
    # Read the data
    for row in sheet.iter_rows(min_row=2):  # Skip header row
        track_name = row[0].value
        coord_type = row[1].value
        
        if track_name not in tracks_data:
            tracks_data[track_name] = {'X': [], 'Y': [], 'frames': []}
        
        # Get coordinates for this track
        coords = [cell.value for cell in row[2:]]
        tracks_data[track_name][coord_type] = coords
        
        # Store frame numbers for this track
        if coord_type == 'X':  # Only store frames once per track
            tracks_data[track_name]['frames'] = frame_numbers
    
    # Create tracks in Blender
    for track_name, data in tracks_data.items():
        # Create new track
        track = clip.tracking.tracks.new(name=track_name)
        
        # Add markers for each frame
        for i, (x, y) in enumerate(zip(data['X'], data['Y'])):
            if x is not None and y is not None:
                # Convert coordinates from pixel space to normalized space (0-1)
                normalized_x = x / video_width
                normalized_y = y / video_height
                
                # Create marker at the current frame using the exact frame number
                marker = track.markers.insert_frame(data['frames'][i])
                marker.co = (normalized_x, normalized_y)

class SaveDirectory(PropertyGroup):
    path: StringProperty(
        name="",
        description="Path to Excel File",
        default="",
        maxlen=1024,
        subtype='FILE_PATH')

class CLIP_PT_import_track(bpy.types.Panel):
    bl_space_type = 'CLIP_EDITOR'
    bl_region_type = 'UI'
    bl_category = "Track"
    bl_label = "Import Track"
    bl_options = {'DEFAULT_CLOSED'}

    def draw(self, context):
        layout = self.layout
        layout.use_property_split = True
        layout.use_property_decorate = False

        self.layout.label(text="Select Excel File")

        scn = context.scene
        col = layout.column(align=True)
        col.prop(scn.my_tool, "path", text="")
        
        clip = context.space_data.clip
        if not clip:
            layout.active = False
            layout.label(text="No active clip")
            return

        layout.operator("wm.import_all_tracks")

class ImportAllTracks(bpy.types.Operator):
    bl_idname = "wm.import_all_tracks"
    bl_label = "Import all Tracks"

    def execute(self, context):
        file_path = context.scene.my_tool.path
        
        # Check if the file exists and is an Excel file
        if not file_path or not os.path.isfile(file_path):
            self.report({"ERROR"}, "Please select a valid Excel file.")
            return {'CANCELLED'}
            
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            self.report({"ERROR"}, "Selected file is not an Excel file.")
            return {'CANCELLED'}
        
        clip = context.space_data.clip
        if not clip:
            self.report({"ERROR"}, "No active clip.")
            return {'CANCELLED'}
            
        try:
            import_tracks_from_excel(file_path, clip)
            self.report({'INFO'}, f"Imported all tracks into current clip: {clip.name}")
        except Exception as e:
            self.report({"ERROR"}, f"Error importing tracks: {str(e)}")
            return {'CANCELLED'}

        return {'FINISHED'}

classes = (
    CLIP_PT_import_track,
    ImportAllTracks,
    SaveDirectory
)

def register():
    for cls in classes:
        bpy.utils.register_class(cls)
    bpy.types.Scene.my_tool = PointerProperty(type=SaveDirectory)

def unregister():
    for cls in classes:
        bpy.utils.unregister_class(cls)
    del bpy.types.Scene.my_tool

if __name__ == "__main__":
    register() 