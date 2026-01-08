## HOW TO USE
# 1) In Object mode click object -> Object -> Set Origin -> CoM (Volume)
# 2) Go to edit mode
# 3) Shift and left click and select TWO vertices
# 4) Right click in 3D viewpoint and select "Export Centroid and Vertex Coordinates"
# 5) CoM = 0, 0, 0 rest are valid coordinates


bl_info = {
    "name": "Export Centroid and Vertex Coordinates",
    "author": "Viggo",
    "version": (1, 1),
    "blender": (3, 0, 0),
    "location": "Mesh > Context Menu",
    "description": "Export the centroid and coordinates of two selected vertices to an Excel file",
    "category": "Mesh",
}

import bpy
import mathutils
import bmesh  # Import the bmesh module for working with Edit Mode meshes
from openpyxl import Workbook, load_workbook
import os

class MESH_OT_ExportCentroidAndVertices(bpy.types.Operator):
    """Export the centroid and coordinates of two selected vertices"""
    bl_idname = "mesh.export_centroid_vertices"
    bl_label = "Export Centroid and Vertices"
    bl_options = {'REGISTER', 'UNDO'}
    
    def execute(self, context):
        obj = context.object

        # Check if an object is selected and is a mesh
        if not obj or obj.type != 'MESH':
            self.report({'ERROR'}, "No mesh object selected.")
            return {'CANCELLED'}

        # Get selected vertices in Edit Mode
        bm = bmesh.from_edit_mesh(obj.data)
        selected_vertices = [v for v in bm.verts if v.select]

        if len(selected_vertices) != 2:
            self.report({'ERROR'}, "Exactly two vertices must be selected.")
            return {'CANCELLED'}

        try:
            # Try to get vertices in selection order
            bm.select_history.validate()
            
            if len(bm.select_history) >= 2:
                # Get the last two selected vertices in order
                last_selected = [elem for elem in bm.select_history if isinstance(elem, bmesh.types.BMVert)]
                
                if len(last_selected) >= 2:
                    vertex_1 = last_selected[-2].co  # First selected (second-to-last in history)
                    vertex_2 = last_selected[-1].co  # Second selected (last in history)
                    print("Using selection order from history")
                else:
                    # Fallback to index order
                    vertex_1 = selected_vertices[0].co
                    vertex_2 = selected_vertices[1].co
                    print(f"Using index order - Vertex 1 index: {selected_vertices[0].index}, Vertex 2 index: {selected_vertices[1].index}")
            else:
                # Fallback to index order
                vertex_1 = selected_vertices[0].co
                vertex_2 = selected_vertices[1].co
                print(f"Using index order - Vertex 1 index: {selected_vertices[0].index}, Vertex 2 index: {selected_vertices[1].index}")

            # Calculate the centroid
            vertices = [v.co for v in obj.data.vertices]
            num_vertices = len(vertices)

            if num_vertices == 0:
                self.report({'ERROR'}, "Mesh has no vertices.")
                return {'CANCELLED'}

            centroid = mathutils.Vector((0, 0, 0))
            for vertex in vertices:
                centroid += vertex
            centroid /= num_vertices

            # Output results to the console
            print(f"Centroid: {centroid}")
            print(f"Vertex 1: {vertex_1}")
            print(f"Vertex 2: {vertex_2}")

            # Save data to an Excel file
            directory = "/Users/viggorey/Desktop/PhD/Cambridge/Macaranga/CoM/"
            if not os.path.exists(directory):
                os.makedirs(directory)

            file_path = os.path.join(directory, "centroid_vertex_data.xlsx")

            if not os.path.exists(file_path):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Centroid and Vertex Data"
                sheet.append(["Object Name", "Centroid X", "Centroid Y", "Centroid Z",
                              "Vertex 1 X", "Vertex 1 Y", "Vertex 1 Z",
                              "Vertex 2 X", "Vertex 2 Y", "Vertex 2 Z"])
                workbook.save(file_path)
            else:
                workbook = load_workbook(file_path)
                sheet = workbook.active

            # Append data
            sheet.append([
                obj.name,
                centroid.x, centroid.y, centroid.z,
                vertex_1.x, vertex_1.y, vertex_1.z,
                vertex_2.x, vertex_2.y, vertex_2.z
            ])
            workbook.save(file_path)

            self.report({'INFO'}, f"Centroid and vertex data saved to {file_path}")

            return {'FINISHED'}

        except Exception as e:
            self.report({'ERROR'}, f"Failed to export data: {e}")
            return {'CANCELLED'}

# Register the operator
def menu_func(self, context):
    self.layout.operator(MESH_OT_ExportCentroidAndVertices.bl_idname)

def register():
    bpy.utils.register_class(MESH_OT_ExportCentroidAndVertices)
    bpy.types.VIEW3D_MT_edit_mesh_context_menu.append(menu_func)

def unregister():
    bpy.utils.unregister_class(MESH_OT_ExportCentroidAndVertices)
    bpy.types.VIEW3D_MT_edit_mesh_context_menu.remove(menu_func)

if __name__ == "__main__":
    register()