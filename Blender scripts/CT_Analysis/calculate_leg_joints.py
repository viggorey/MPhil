## HOW TO USE
# 1) Go to edit mode
# 2) Shift and left click to select vertices in the order you want them exported
# 3) Right click in 3D viewport and select "Export Vertex Positions"
# 4) Vertices will be exported to Excel in the order you selected them
# 5) IN TRIM CODE ITS FRONT POINTS TO BACK ON ONE SIDE (point 1 is point 2 and point 5 is point 3)

bl_info = {
    "name": "Export Vertex Positions in Order",
    "author": "Viggo",
    "version": (1, 0),
    "blender": (3, 0, 0),
    "location": "Mesh > Context Menu",
    "description": "Export the coordinates of selected vertices to an Excel file in selection order",
    "category": "Mesh",
}

import bpy
import bmesh
from openpyxl import Workbook, load_workbook
import os

class MESH_OT_ExportVertexPositions(bpy.types.Operator):
    """Export the coordinates of selected vertices in selection order"""
    bl_idname = "mesh.export_vertex_positions"
    bl_label = "Export Vertex Positions"
    bl_options = {'REGISTER', 'UNDO'}
    
    def execute(self, context):
        obj = context.object

        # Check if an object is selected and is a mesh
        if not obj or obj.type != 'MESH':
            self.report({'ERROR'}, "No mesh object selected.")
            return {'CANCELLED'}

        # Get selected vertices in Edit Mode
        bm = bmesh.from_edit_mesh(obj.data)
        selected_vertices = [v for v in bm.verts if v.select]

        # Check if any vertices are selected
        if len(selected_vertices) == 0:
            self.report({'WARNING'}, "No vertices selected.")
            return {'CANCELLED'}

        # Check if too many vertices are selected
        if len(selected_vertices) > 100:
            self.report({'WARNING'}, f"Too many vertices selected ({len(selected_vertices)}). Maximum is 100.")
            return {'CANCELLED'}

        try:
            # Try to get vertices in selection order
            bm.select_history.validate()
            
            vertex_coords = []
            
            if len(bm.select_history) > 0:
                # Get selected vertices from history in order
                selected_in_order = [elem for elem in bm.select_history if isinstance(elem, bmesh.types.BMVert) and elem.select]
                
                if len(selected_in_order) == len(selected_vertices):
                    # We have complete selection history
                    for vertex in selected_in_order:
                        vertex_coords.append(vertex.co.copy())
                    print(f"Using selection order from history - {len(vertex_coords)} vertices")
                else:
                    # Incomplete history, show warning
                    self.report({'WARNING'}, "Selection history incomplete. Cannot determine selection order.")
                    return {'CANCELLED'}
            else:
                # No selection history available
                self.report({'WARNING'}, "No selection history available. Cannot determine selection order.")
                return {'CANCELLED'}

            # Output results to the console
            print(f"Exporting {len(vertex_coords)} vertices:")
            for i, coord in enumerate(vertex_coords):
                print(f"Vertex {i+1}: {coord}")

            # Save data to an Excel file
            directory = "/Users/viggorey/Desktop/PhD/Cambridge/Macaranga/CoM/"
            if not os.path.exists(directory):
                os.makedirs(directory)

            file_path = os.path.join(directory, "Leg_joints.xlsx")

            if not os.path.exists(file_path):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Vertex Positions"
                sheet.append(["Vertex Order", "X", "Y", "Z"])
                workbook.save(file_path)
            else:
                workbook = load_workbook(file_path)
                sheet = workbook.active

            # Append data for each vertex
            for i, coord in enumerate(vertex_coords):
                sheet.append([
                    i + 1,  # Vertex order (1-based)
                    coord.x,
                    coord.y,
                    coord.z
                ])
            
            workbook.save(file_path)

            self.report({'INFO'}, f"Exported {len(vertex_coords)} vertices to {file_path}")

            return {'FINISHED'}

        except Exception as e:
            self.report({'ERROR'}, f"Failed to export data: {e}")
            return {'CANCELLED'}

# Register the operator
def menu_func(self, context):
    self.layout.operator(MESH_OT_ExportVertexPositions.bl_idname)

def register():
    bpy.utils.register_class(MESH_OT_ExportVertexPositions)
    bpy.types.VIEW3D_MT_edit_mesh_context_menu.append(menu_func)

def unregister():
    bpy.utils.unregister_class(MESH_OT_ExportVertexPositions)
    bpy.types.VIEW3D_MT_edit_mesh_context_menu.remove(menu_func)

if __name__ == "__main__":
    register()